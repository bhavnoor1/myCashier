import tkinter as tk
from TKclear import TKclear
import tkinter.messagebox as wmsg
import settingsMenu as settings
import filemanager
from datetime import date
#import os
import tkinter.messagebox as rnuo
from ppcalc import ppcalc
from languagemanager import lm, ApplyLanguage
from settingsMenu import GetCurrentTheme


"""hello"""

finishedflag = False

#CurrentRegister = None
gwindow = None
gframe  = None

def QuitToMainMenu(skipCondition = False):
	from mainmenu import mainmain
#if skipcondition is true we dont ask the user to confirm. this is used when clicking on finish
	if skipCondition or wmsg.askquestion(message=lm.GetVar("REG_QUIT")) == "yes":
		TKclear(gframe)
		mainmain(gwindow, gframe)
#END

#checks to make sure the input enterd is valid (like the fields are numbers etc\)
def ValidateInput():
	
	"""checks for invalid input in the fields, and turns all fields with bad input to be red\n
			Returns: True if all fields are correct, False otherwise"""

	theme = GetCurrentTheme()
	incorrectBG = theme.EntryFieldTheming.incorrectBg
	correctBG = theme.EntryFieldTheming.background

	global coinsfields, notesfields, pcardsfields, tillreadfields, vouchersfields, rollsfields
	
	#this flag will be returned. true is "no problems", false is "there are problems"
	flag = True

	for fields in coinsfields + notesfields:
		for field in fields:
			try:
				int(field.get())
				field["bg"] = correctBG  #if everything went correctly set the colour to white
			except ValueError:
				field["bg"] = incorrectBG    #if the input was not valid (ValueError) set it to red
				flag = False
	#END FOR
	
	for pcardsfield in pcardsfields:
		try:
			#checks two things: if it is a float and if it has less than or equal to two decimal places
			#if either of the checks fail we have a bad input
			if NumberHasDecimalPlaces(float(pcardsfield.get()), 2) is not True: raise 69
			pcardsfield["bg"] = correctBG
		except:
			flag = False
			pcardsfield["bg"] = incorrectBG
	#END FOR
	
	for tillreadfield in tillreadfields:
		try:
			if NumberHasDecimalPlaces(float(tillreadfield.get()), 2) is not True: raise 69
			tillreadfield["bg"] = correctBG
		except:
			flag = False
			tillreadfield["bg"] = incorrectBG
	#END FOR
	
	for vouchersfield in vouchersfields:
		try:
			int(vouchersfield.get())
			vouchersfield["bg"] = correctBG
		except:
			flag = False
			vouchersfield["bg"] = incorrectBG
	#END FOR

	for rollsfield in rollsfields:
		try:
			int(rollsfield.get())
			rollsfield["bg"] = correctBG
		except:
			flag = False
			rollsfield["bg"] = incorrectBG
	#END FOR

	return flag
#END
def Finish():
	"""This function saves all changes to the file and goes back to main menu, if valid"""
	
	try:
		from openpyxl import load_workbook
		import GODPROGRAMMING

		#copy the temp file to a new file in the save data path
		fname = filemanager.GetTodaysFileWithPath()
		GODPROGRAMMING.Pray(fname)		#if you can't understand GODPROGRAMMING, you don't deserve the explanation

		if ValidateInput() is not True:
			return 0
		
		
		wb = load_workbook(filename = fname)
		ss = wb.active

		ss["A1"] = "Date: " + str(date.today())
		floatstring = float(settings.GetSetting("float"))
		floatstring = str(floatstring) + "0" if NumberHasDecimalPlaces(floatstring, 1) else ""
		#in this case the number will be of the form "x.y", so we add a 0 at the end to make it look better
		#if it is not we already have two decimal places so we don't add
		ss["A5"] = ss["E5"] = ss["I5"] = "float: $" + floatstring

		if checkboxes[0].get() == 1:
			ss["C19"] = float(settings.GetSetting("float"))
		if checkboxes[1].get() == 1:
			ss["G19"] = float(settings.GetSetting("float"))

		for Register in [0, 1, 2]:
			regCol = ["A", "E", "I"][Register]  #column to write to

			#if this is true we skip this register because it has been marked as not active
			thirdcol = chr(ord(regCol) + 2) #third column, for float
			if checkboxes[Register].get() == 0:
				ss[regCol + "3"] = "(not active)"
				ss[thirdcol + "19"] = ""
				continue
			else:
				ss[thirdcol + "19"] = float(settings.GetSetting("float")) #set the float


			for i in range(len(coinsfields[Register])):
				ss[regCol + str(6 + i)]  = int(coinsfields[Register][i].get())
			for i in range(len(notesfields[Register])):
				ss[regCol + str(13 + i)] = int(notesfields[Register][i].get())
			
			#some items are in the second column, so do this
			regCol = chr(ord(regCol) + 1)

			ss[regCol + "24"] = round(float(pcardsfields[Register].get()), 2)
			ss[regCol + "26"] = round(float(tillreadfields[Register].get()), 2)

			ss[regCol + "12"] = int(rollsfields[Register].get())
			ss["B" + str(33 + Register)] = int(vouchersfields[Register].get())
		#END FOR
		
		wb.save(fname)
	except Exception as e:
		print(e)
		rnuo.showerror(message=lm.GetVar("REG_SAVE_ERROR"))
		return 0
	
	QuitToMainMenu(True)
#END

#useful in checking for good or bad values
def NumberHasDecimalPlaces(number : float, places : int):
	"""returns true if the given number does not have more than the given number of decimal places"""
	#yes, technically you could enter "inf" in the field and it would accept it
	if number == float("inf") or number == float("-inf") or number == float("nan") or number == float("-nan"):
		return False
	return round(number, places) == number

def PunchTheCards(register : int):
	print(register)
	#the user should be able to enter any amount of punchcards values
	#after finishing, automatically store it in the field. you can access them using the "pcardsfields" list
	#print("Noor do this")
	#rnuo.showerror(message="Noor do this")
	#rnuo.showerror(message="just add them manually dumbass"); return
#	from ppcalc import value
	ppcalc(register)

#this shit fails if you first open settings
def ToggleRegister(a, b, c):
	i = int(a[6:])  #get the number of the barialbe
	i -= int(str(checkboxes[0])[6:])   #if variable of index 0 is called with the number '5', we want to subtract five from it and therefore from all
	print(a, checkboxes[0])

	print("in toggleregister:", i)
	val = checkboxes[i].get()
	print(val)
	state = "disabled" if val == 0 else "normal"
	if val == 0:
		#italian keyboards be like:
		for çéèòàù in coinsfields[i]:
			çéèòàù.delete("0", "end")
			çéèòàù.insert("0", "0")
		for çéèòàù in notesfields[i]:
			çéèòàù.delete("0", "end")
			çéèòàù.insert("0", "0")
		#pcarsd
		pcardsfields[i].delete("0", "end")
		pcardsfields[i].insert("0", "0.00")
		#tillread
		tillreadfields[i].delete("0", "end")
		tillreadfields[i].insert("0", "0.00")
		#vouchers
		vouchersfields[i].delete("0", "end")
		vouchersfields[i].insert("0", "0")
		#rolls
		rollsfields[i].delete("0", "end")
		rollsfields[i].insert("0", "0")
	#we only reset stuff if yes

	#sret steate
	pcardsbtns[i]["state"] = state
	pcardsfields[i]["state"] = state
	tillreadfields[i]["state"] = state
	vouchersfields[i]["state"] = state
	rollsfields[i]["state"] = state
	for çéèòàù in coinsfields[i] + notesfields[i]:
		çéèòàù["state"] = state



totfields = []		#list of fields for totals
def UpdateTotals():
	global finishedflag
	if finishedflag is not True: return		#to avoid index errors because we do it before we place all of them

	theme = GetCurrentTheme()
	incorrectBG = theme.EntryFieldTheming.incorrectBg
	correctBG = theme.EntryFieldTheming.background

	#because i don't want to make things too complicated, this updates every total, every time a change is made
	print("updating totals...")
	global totfields, coinsfields, notesfields		#pp  #vouch   #rolls
	mults = [ .05, .1, .25, 1, 2, 5, 10, 25, 50 ] + [1] + [5]   + [10]
	for i in [0, 1, 2]:
		tot = 0
		combinedlist = (coinsfields[i] + notesfields[i] + [pcardsfields[i]] + [vouchersfields[i]] + [rollsfields[i]])
		for fieldaddr in range(len(combinedlist)):
			singlefield = combinedlist[fieldaddr]
			try:
				delta = round((float(singlefield.get()) * mults[fieldaddr]), 2)
				tot += delta	#we try adding to the total
				singlefield["bg"] = correctBG
			except:		#if it's not a valid value skip it
				singlefield["bg"] = incorrectBG
		#when we reach this we're done calculating the total for each register
		#so we just set the total
		if NumberHasDecimalPlaces(tot, 1):
			tot = str(tot) + "0"
		else:
			tot = str(tot)
		tot = tot[:tot.index(".") + 3]
		totfields[i]["state"] = "normal"
		totfields[i].delete("0", "end")
		totfields[i].insert("0", (tot))
		totfields[i]["state"] = "disabled"
#END FUNC


OpenedOnce = False
def registersmain(window, frame):
	global gwindow, gframe
	from openpyxl import load_workbook

	try:   #if any error happens we'll go back to mainmenu to avoid catastrophes
		"""
		import datetime
		global OpenedOnce
		if datetime.date.today().month == 4 and datetime.date.today().day == 1 and OpenedOnce is False:
			OpenedOnce = True
			#raise rickrollException"""

		print("registers init")
		print("initialising variables...")
		
		print("initialising global variables...")
		global coinsfields, notesfields #these will be lists of lists
		coinsfields, notesfields = [], []
		global pcardsfields, tillreadfields, vouchersfields, rollsfields
		pcardsfields, tillreadfields, vouchersfields, rollsfields = [], [], [], []
		global checkboxes, pcardsbtns
		pcardsbtns = []
		
		try:
			len(checkboxes)  #we try to use checkboxes, to trigger an exception if it's not initialised
			hadVars = True  #if we don't get an exception we have vars
		except:
			#if it is not we initialise it
			checkboxes = []
			hadVars = False

		#we don't initialise checkboxes because we reuse variables

		#if we changed the window we want to reset the variables as the variables depend on the windows
		if (window, frame) != (gwindow, gframe):
			checkboxes = []
			hadVars = False

		if window == None or frame == None:
			window, frame = gwindow, gframe
		else:
			gwindow, gframe = window, frame

		print("placing static window elements...")
		window.title(lm.GetVar("REG_TITLE"))

		COINSOFFSET = 5
		ROLLSOFFSET = COINSOFFSET + 5
		THIRDOFFSET = ROLLSOFFSET + 9
		COLUMN = 2

		tk.Label(frame, text="$0.05") .grid(row=COINSOFFSET+1,column=COLUMN)
		tk.Label(frame, text="$0.10").grid(row=COINSOFFSET+2,column=COLUMN)
		tk.Label(frame, text="$0.25").grid(row=COINSOFFSET+3,column=COLUMN)
		tk.Label(frame, text="$1") .grid(row=COINSOFFSET+4,column=COLUMN)
		tk.Label(frame, text="$2") .grid(row=COINSOFFSET+5,column=COLUMN)
		
		tk.Label(frame, text="$5") .grid(row=ROLLSOFFSET+1,column=COLUMN)
		tk.Label(frame, text="$10").grid(row=ROLLSOFFSET+2,column=COLUMN)
		tk.Label(frame, text="$25").grid(row=ROLLSOFFSET+3,column=COLUMN)
		tk.Label(frame, text="$50").grid(row=ROLLSOFFSET+4,column=COLUMN)

		tk.Label(frame, text="REG_PCARDS").grid(row=THIRDOFFSET+0, column=COLUMN)
		tk.Label(frame, text="REG_TILL").grid(row=THIRDOFFSET+1, column=COLUMN)
		tk.Label(frame, text="REG_VOUCHERS").grid(row=THIRDOFFSET+2, column=COLUMN)

		tk.Label(frame, text="REG_ROLLS").grid(row=THIRDOFFSET+3, column=COLUMN)

		try:
			wb = load_workbook(filename = filemanager.GetTodaysFileWithPath()).active
		except:
			wb = None
		regNumToColumn = ["A", "E", "I"]  #columns in the spreadsheet
		
		tk.Label(frame, width=15, text=lm.GetVar("REG_FLOAT") + str(settings.GetSetting("float"))).grid(column=2, row=0)

		#stuff of the punch carsds
		#for some reason we have to define it outside of the loop otherwise it doesnt work.
		#but its easier to grid it in the loop, so here we are: beautiful, absolutely not confusing code
		pcardsbtns.append(
			tk.Button(frame, padx=14, text="REG_PP_A", command=lambda:PunchTheCards(0), width=4)
		)
		pcardsbtns.append(
			tk.Button(frame, padx=14, text="REG_PP_B", command=lambda:PunchTheCards(1), width=4)
		)
		pcardsbtns.append(
			tk.Button(frame, padx=14, text="REG_PP_C", command=lambda:PunchTheCards(2), width=4)
		)


		COLOFFSET = 4  #COLumn OFFSET
		for CurrentRegister in [0,1,2]:
			print("initialising constants and variables for register", str(CurrentRegister) + "...")
			fstCol = regNumToColumn[CurrentRegister] 	#first column of the register
			secCol = chr(1 + ord(fstCol))				#second column of the register
			col = COLOFFSET + 2 * CurrentRegister + 1	#column the fields will be placed on
							#2 because we'll have a separator

			coinsfields.append([]); notesfields.append([])
			
			print("placing text fields...")
			for i in range(1, 6):  #(A/E/I)6-10
				temp = tk.Entry(frame, width=7)#, validate="focus", validatecommand=UpdateTotals)
				temp.grid(column=col, row=COINSOFFSET + i)
				try:
					temp.insert("end",
					#this just means the following:
					#if there is no file, the default text is 0.
					#if there is a file but it has nothing in it, the default text is 0.
					#in all other cases, default text is what you see in the file
						"0" if wb is None
							else int(wb[fstCol + str(5 + i)].value or 0)
					)
				except Exception as e:
					print(e)
					temp.insert("end", "") #this is triggered for example if text data is found
					#the only way this could happen is if they change the file manually, which should not happen, but this way we avoid unnecessary exceptions
				#once we're done with all the operations add to the list
				coinsfields[CurrentRegister].append(temp)
			#END FOR
			for i in range(1, 5):  #(A/E/I)13-16
				temp = tk.Entry(frame, width=7)#, validate="focus", validatecommand=UpdateTotals)
				temp.grid(column=col, row=ROLLSOFFSET + i)
				try:
					temp.insert("end",
						"0" if wb is None
							else int(wb[fstCol + str(12 + i)].value or 0)
					)
				except Exception as e:
					print(e)
					temp.insert("end", "")
				notesfields[CurrentRegister].append(temp)
			#END FOR

			pcardsfields.append(tk.Entry(frame, width=7))
			pcardsfields[-1].grid(row=THIRDOFFSET+0, column=col)
			try:
				pcardsfields[-1].insert("end",
					"0.0" if wb is None
						else round(
							float(wb[secCol + "24"].value or "0"),  #if the value is none set to 0
							2) #round to 2 decimals
				)
				print(pcardsfields[-1].get())
				#to make it look better money will always have two decimal places
				#floats ave always at least one decimal places when printed, so we only need to check this
				if NumberHasDecimalPlaces(float(pcardsfields[-1].get()), 1): pcardsfields[-1].insert("end", "0")
			except Exception as e:
				print(e)
				tillreadfields[-1].delete("0", "end")
				pcardsfields[-1].insert("end", "0.00")

			tillreadfields.append(tk.Entry(frame, width=7))
			tillreadfields[-1].grid(row=THIRDOFFSET+1, column=col)
			try:
				tillreadfields[-1].insert("end",
					"0.0" if wb is None
						else round(
							float(wb[secCol + "26"].value or 0),  #if the value is none set to 0
							2) #round to 2 decimals
				)
				if NumberHasDecimalPlaces(float(tillreadfields[-1].get()), 1): tillreadfields[-1].insert("end", "0")
			except Exception as e:
				print(e)
				tillreadfields[-1].delete("0", "end")
				tillreadfields[-1].insert("end", "0.00")
			
			vouchersfields.append(tk.Entry(frame, width=7))
			vouchersfields[-1].grid(row=THIRDOFFSET+2, column=col)
			try:
				vouchersfields[-1].insert("end",
					"0" if wb is None
						else wb["B" + str(33 + CurrentRegister)].value or 0  #if the value is none set to 0
				)
				print("vouchers: " + vouchersfields[-1].get())
			except Exception as e:
				print(e)
				vouchersfields[-1].insert("end", "0")

			rollsfields.append(tk.Entry(frame, width=7))
			rollsfields[-1].grid(row=THIRDOFFSET+3, column=col)
			try:
				rollsfields[-1].insert("end",
					"0" if wb is None
						else wb[secCol + "12"].value or 0  #if the value is none set to 0
				)
				print("vouchers: " + rollsfields[-1].get())
			except Exception as e:
				print(e)
				rollsfields[-1].insert("end", "0")


			"""#txtcolumn = ["C", "G", "K"][CurrentRegister]
			#now we trigger the valuechange event if the register has to be disabled
			#var.set(1 if wb is None else int(bool(wb[txtcolumn + "19"].value)))
			#if we don't have a value we default to active
			#if we do have a value and it's null we set to disable, else enable
"""
			print("adding buttons...")
			pcardsbtns[CurrentRegister].grid(row=THIRDOFFSET-1, column=col)

			print("adding labels...")
			tk.Label(frame, text= lm.GetVar("REG_NAME") + " " + chr(0x41 + CurrentRegister), font='Helvetica 10 bold').grid(row=COINSOFFSET, column=col)
			tk.Label(frame, width=5).grid(column=col-1)
		#END REGISTER FOR

		for CurrentRegister in [0, 1, 2]:
			fstCol = regNumToColumn[CurrentRegister] 	#first column of the register
			secCol = chr(1 + ord(fstCol))               #second column of the register
			col = COLOFFSET + 2 * CurrentRegister + 1	#column the fields will be placed on

			#its important we do it after everything else bc otherwise we'll reference undefined widgets
			#also, it's in a separate loop bc this way we can create the variables one after the other which makes it easier to get their number
			if hadVars is False: #we want to create new vars
				var = tk.IntVar(frame, 1)
				checkboxes.append(var)
			else:	 			 #we reuse the old one
				var = checkboxes[CurrentRegister]
			var.trace_add("write", ToggleRegister)  #on write to the var this func will be called
			tk.Checkbutton(frame, text="REG_ACT", variable=var).grid(column=col, row=COINSOFFSET-1)

		print("placing buttons...")
		#tk.Button(frame, text="save", width=5, command=SaveRegister).grid(row=17,column=1)

		tk.Button(frame, text="REG_BACK", width=5, command=QuitToMainMenu).grid(row=THIRDOFFSET+6,column=COLUMN)
		tk.Button(frame, text="REG_SAVE", width=5, command=Finish).grid(row=THIRDOFFSET+5,column=COLUMN)

		print("placing totals...")
		global totfields
		totfields.clear()	#this way we won't be referencing previously defined labelese
		tk.Label(frame, text=lm.GetVar("REG_TOTAL_TITLE") + " A").grid(row=7,  column=11)
		totfields.append(
			tk.Entry(frame, width=20, justify="right", state="disabled")
		)
		tk.Label(frame, text=lm.GetVar("REG_TOTAL_TITLE") + " B").grid(row=10, column=11)
		totfields.append(
			tk.Entry(frame, width=20, justify="right", state="disabled")
		)
		tk.Label(frame, text=lm.GetVar("REG_TOTAL_TITLE") + " C").grid(row=13, column=11)
		totfields.append(
			tk.Entry(frame, width=20, justify="right", state="disabled")
		)
		totfields[0].grid(row=8,  column=11)
		totfields[1].grid(row=11, column=11)
		totfields[2].grid(row=14, column=11)

		tk.Button(frame, text=lm.GetVar("REG_TOT_UPDATE"), command=UpdateTotals).grid(row=THIRDOFFSET-1, column=11)

		print("placing the empty labels...")
		#empty labels for spaces and style
		tk.Label(frame).grid(row=ROLLSOFFSET-1)
		tk.Label(frame).grid(row=ROLLSOFFSET+5)
		tk.Label(frame).grid(row=THIRDOFFSET+3)
		tk.Label(frame).grid(row=THIRDOFFSET+4)
		tk.Label(frame, width=5).grid(column=3)
		tk.Label(frame, width=5).grid(column=10)

		settings.ApplyTheme(window, frame)
		ApplyLanguage(frame)
	
		#used by the update totals function
		global finishedflag
		finishedflag = True
		UpdateTotals()			#this way we can show what the actual totals are

		window.mainloop()

	except Exception as e:
		import mainmenu
		import errormanager
		errormanager.PrintErrorLog(e)
		print(e)
		TKclear(frame)
		rnuo.showerror(message=lm.GetVar("REG_CRASH"))
		#try:
		#	os.remove(filemanager.GetTodaysFileWithPath())  #the error was most likely caused by this, so remove it
		#except: pass
		mainmenu.mainmain(window, frame)
		